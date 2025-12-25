import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "AmazonCable成本核算"

# --- 定义样式 ---
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
input_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
result_font = Font(bold=True)
result_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

# --- 数据结构与公式 (已修正行号引用) ---
# Excel行号对照：
# Row 1: 标题
# Row 2: 铜价输入 (B2)
# Row 3: 料价输入 (B3)
# Row 4: 铜价转换 (B4)
# ...
# Row 25: 铜成本 (B25)
# Row 26: 塑料成本 (B26)
# Row 27: 制造费 (B27)
# Row 28: 直接材料合计 (B28) = B25+B26
# Row 29: 出厂总成本 (B29) = B28+B27
# Row 31: 利润率 (B31)
# Row 32: 最终报价 (B32) = B29*(1+B31)

data = [
    # [项目, 数值/公式, 单位, 说明, 样式类型]
    ["一、原材料市场价格", "", "", "", "title"],
    ["1#电解铜价格", 12200, "USD/吨", "请输入今日铜价", "input"],
    ["绝缘/护套料均价", 0.85, "USD/kg", "综合料价", "input"],
    ["铜价(转换后)", "=B2/1000", "USD/kg", "自动计算", "calc"],
    
    ["二、电缆规格参数 (YJV 4*70+1*35)", "", "", "", "title"],
    ["主线芯截面", 70, "mm²", "主导", "input"],
    ["主线芯数量", 4, "根", "", "input"],
    ["中性线芯截面", 35, "mm²", "地线/零线", "input"],
    ["中性线芯数量", 1, "根", "", "input"],
    ["绝缘厚度(主)", 1.1, "mm", "查表7", "input"],
    ["绝缘厚度(副)", 0.9, "mm", "查表7", "input"],
    ["成缆后总外径", 33.0, "mm", "查系数K", "input"],
    ["外护套厚度", 1.8, "mm", "国标", "input"],
    
    ["三、密度常数", "", "", "", "title"],
    ["铜密度", 8.89, "g/cm³", "", "const"],
    ["绝缘料密度(XLPE)", 0.92, "g/cm³", "", "const"],
    ["护套料密度(PVC)", 1.45, "g/cm³", "", "const"],
    ["填充系数", 0.3, "系数", "估算值", "const"],

    ["四、用量自动计算 (kg/km)", "", "", "", "title"],
    # B20: 铜重 = (主截面*主数 + 副截面*副数) * 铜密度
    ["1. 铜导体总重", "=(B6*B7+B8*B9)*B15", "kg/km", "导体重量", "calc"],
    # B21: 绝缘重 (几何公式)
    ["2. 绝缘料总重", "=(B7*3.14*((2*SQRT(B6/3.14)+2*B10)^2-(2*SQRT(B6/3.14))^2)/4 + B9*3.14*((2*SQRT(B8/3.14)+2*B11)^2-(2*SQRT(B8/3.14))^2)/4)*B16", "kg/km", "绝缘体积x密度", "calc"],
    # B22: 护套重
    ["3. 护套料总重", "=3.14*((B12+2*B13)^2-B12^2)/4*B17", "kg/km", "护套体积x密度", "calc"],
    # B23: 填充重
    ["4. 填充及其他", "=(B21+B22)*B18", "kg/km", "按塑料重30%估算", "calc"],
    
    ["五、成本分析 (USD/m)", "", "", "", "title"],
    # B25: 铜成本 = 铜重(B20) * 铜价(B4) / 1000
    ["铜材成本", "=B20*B4/1000", "USD/m", "核心成本", "result"],
    # B26: 塑料成本 = (绝缘+护套+填充) * 料价(B3) / 1000
    ["塑料及辅材成本", "=(B21+B22+B23)*B3/1000", "USD/m", "辅材成本", "result"],
    # B27: 制造费 (输入)
    ["制造费用(水电人工)", 0.45, "USD/m", "工厂分摊", "input"],
    # B28: 直接材料 = 铜(B25) + 塑(B26)
    ["直接材料成本", "=B25+B26", "USD/m", "光料成本", "result"],
    # B29: 总成本 = 材料(B28) + 制造(B27)
    ["出厂总成本", "=B28+B27", "USD/m", "保本价", "result"],
    
    ["六、报价建议", "", "", "", "title"],
    # B31: 利润率
    ["目标利润率", 0.10, "%", "10%=0.1", "input"],
    # B32: 最终价 = 总成本(B29) * (1+利润率(B31))
    ["建议报价(含税)", "=B29*(1+B31)", "USD/m", "最终价格", "result_final"]
]

# --- 写入数据 ---
for row_idx, row_data in enumerate(data, 1):
    ws.cell(row=row_idx, column=1, value=row_data[0])
    ws.cell(row=row_idx, column=2, value=row_data[1])
    ws.cell(row=row_idx, column=3, value=row_data[2])
    ws.cell(row=row_idx, column=4, value=row_data[3])
    
    row_type = row_data[4]
    
    # 边框
    for col in range(1, 5):
        ws.cell(row=row_idx, column=col).border = thin_border
        
    # 样式逻辑
    if row_type == "title":
        for col in range(1, 5):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = header_font
            cell.fill = header_fill
    elif row_type == "input":
        ws.cell(row=row_idx, column=2).fill = input_fill
    elif "result" in row_type:
        ws.cell(row=row_idx, column=2).font = result_font
        if row_type == "result_final":
            ws.cell(row=row_idx, column=2).fill = result_fill
            ws.cell(row=row_idx, column=2).font = Font(bold=True, size=12, color="FF0000")

# --- 列宽调整 ---
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 20

# 保存
file_name = "AmazonCable_Cost_Calculator_v2.xlsx"
wb.save(file_name)
print(f"已生成修正版表格：{file_name}")
print("公式逻辑已重新校对：")
print("1. 直接材料成本 = 铜成本(B25) + 塑料成本(B26)")
print("2. 出厂总成本 = 直接材料(B28) + 制造费(B27)")
print("3. 最终报价 = 总成本(B29) * (1 + 利润率B31)")